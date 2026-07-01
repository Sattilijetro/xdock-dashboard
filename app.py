import streamlit as st
import pandas as pd
import io
import re
import sys
import tempfile
from copy import copy
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
try:
    from validate_invoice import run as validate_run
    FREEZPAK_ENABLED = True
except ImportError:
    FREEZPAK_ENABLED = False

st.set_page_config(
    page_title="Invoice Automation Dashboard",
    page_icon="\U0001f4e6",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');
html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; background-color: #0f1117; color: #e0e0e0; }
.dashboard-header { background: linear-gradient(135deg, #1a1f2e 0%, #16213e 100%); border: 1px solid #2a3a5c; border-radius: 12px; padding: 28px 36px; margin-bottom: 28px; display: flex; align-items: center; gap: 20px; }
.dashboard-title { font-family: 'IBM Plex Mono', monospace; font-size: 26px; font-weight: 600; color: #4fc3f7; letter-spacing: -0.5px; margin: 0; }
.dashboard-subtitle { font-size: 13px; color: #7a8aaa; margin: 4px 0 0 0; font-weight: 300; }
.status-dot { width: 10px; height: 10px; border-radius: 50%; background: #4caf50; box-shadow: 0 0 8px #4caf50; display: inline-block; margin-right: 8px; }
.step-badge { display: inline-block; background: #4fc3f7; color: #0f1117; font-family: 'IBM Plex Mono', monospace; font-size: 11px; font-weight: 600; padding: 2px 10px; border-radius: 20px; margin-right: 8px; }
.info-box { background: #1a2744; border-left: 3px solid #4fc3f7; border-radius: 0 8px 8px 0; padding: 12px 16px; margin: 10px 0; font-size: 13px; color: #b0c4de; }
.success-box { background: #1a2e1a; border-left: 3px solid #4caf50; border-radius: 0 8px 8px 0; padding: 12px 16px; margin: 10px 0; font-size: 13px; color: #a5d6a7; }
.warning-box { background: #2a2010; border-left: 3px solid #ff9800; border-radius: 0 8px 8px 0; padding: 12px 16px; margin: 10px 0; font-size: 13px; color: #ffcc80; }
.error-box { background: #2a1010; border-left: 3px solid #f44336; border-radius: 0 8px 8px 0; padding: 12px 16px; margin: 10px 0; font-size: 13px; color: #ef9a9a; }
.placeholder-box { background: #1a1a2e; border-left: 3px solid #546e8a; border-radius: 0 8px 8px 0; padding: 12px 16px; margin: 10px 0; font-size: 13px; color: #7a8aaa; }
.metric-row { display: flex; gap: 12px; margin: 16px 0; }
.metric-card { background: #1a1f2e; border: 1px solid #2a3a5c; border-radius: 10px; padding: 16px 20px; flex: 1; text-align: center; }
.metric-value { font-family: 'IBM Plex Mono', monospace; font-size: 22px; font-weight: 600; color: #4fc3f7; }
.metric-label { font-size: 11px; color: #7a8aaa; margin-top: 2px; text-transform: uppercase; letter-spacing: 0.5px; }
.sub-type-label { font-family: 'IBM Plex Mono', monospace; font-size: 11px; color: #7a8aaa; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px; }
.stTabs [data-baseweb="tab-list"] { background: #1a1f2e; border-radius: 10px; padding: 4px; gap: 4px; border: 1px solid #2a3a5c; }
.stTabs [data-baseweb="tab"] { border-radius: 8px; color: #7a8aaa; font-family: 'IBM Plex Mono', monospace; font-size: 13px; padding: 8px 20px; }
.stTabs [aria-selected="true"] { background: #4fc3f7 !important; color: #0f1117 !important; font-weight: 600; }
.stButton > button { background: #4fc3f7; color: #0f1117; border: none; border-radius: 8px; font-family: 'IBM Plex Mono', monospace; font-weight: 600; font-size: 13px; padding: 10px 28px; width: 100%; transition: opacity 0.2s; }
.stButton > button:hover { opacity: 0.85; }
.stDataFrame { border-radius: 10px; border: 1px solid #2a3a5c; overflow: hidden; }
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 24px 32px; }
.mode-toggle-wrap { margin-bottom: 20px; }
.mode-toggle-wrap div[data-testid="stRadio"] > div { display: flex; gap: 0; background: #1a1f2e; border: 1px solid #2a3a5c; border-radius: 10px; padding: 4px; width: fit-content; }
.mode-toggle-wrap div[data-testid="stRadio"] label { cursor: pointer; font-family: 'IBM Plex Mono', monospace; font-size: 13px; font-weight: 600; padding: 8px 22px; border-radius: 7px; color: #7a8aaa; transition: all 0.15s; margin: 0; }
.mode-toggle-wrap div[data-testid="stRadio"] label:has(input:checked) { background: #4fc3f7; color: #0f1117; }
.mode-toggle-wrap div[data-testid="stRadio"] label:hover:not(:has(input:checked)) { background: #232b3e; color: #b0c4de; }
.mode-toggle-wrap div[data-testid="stRadio"] [data-testid="stMarkdownContainer"] p { margin: 0; line-height: 1; }
.mode-toggle-wrap div[data-testid="stRadio"] input[type="radio"] { display: none; }
</style>
""", unsafe_allow_html=True)

st.markdown(f"""
<div class="dashboard-header">
    <div style="font-size:36px">\U0001f4e6</div>
    <div>
        <p class="dashboard-title">XDOCK INVOICE DASHBOARD</p>
        <p class="dashboard-subtitle">
            <span class="status-dot"></span>
            Jetro / Restaurant Depot &nbsp;&#183;&nbsp; Logistics Operations &nbsp;&#183;&nbsp;
            {datetime.now().strftime("%B %d, %Y")}
        </p>
    </div>
</div>
""", unsafe_allow_html=True)

XDOCKS = {
    "\U0001f3ed  Halls":    {"key": "halls",    "desc": "Halls XDock invoices",    "color": "#4fc3f7"},
    "❄️  Freezpak": {"key": "freezpak", "desc": "Freezpak XDock invoices", "color": "#81d4fa"},
    "\U0001f4e6  EXP":      {"key": "exp",      "desc": "EXP XDock invoices",      "color": "#4db6ac"},
    "\U0001f69a  EFH":      {"key": "efh",      "desc": "EFH XDock invoices",      "color": "#80cbc4"},
}

INVOICE_TYPES = {
    "halls": [
        {"name": "IBT",            "key": "halls_ibt",          "placeholder": True},
        {"name": "Trucking & FSC", "key": "halls_trucking_fsc",  "placeholder": False},
        {
            "name": "Warehousing", "key": "halls_warehousing", "placeholder": False,
            "sub_types": [
                {"name": "Ancillary",         "key": "halls_warehousing_ancillary",         "placeholder": False},
                {"name": "Ancillary Invoice", "key": "halls_warehousing_ancillary_invoice",  "placeholder": False},
                {"name": "Inbound",           "key": "halls_warehousing_inbound",            "placeholder": False},
                {"name": "Renewal",           "key": "halls_warehousing_renewal",            "placeholder": False},
                {"name": "Sort & Selection",  "key": "halls_warehousing_sort_selection",     "placeholder": False},
            ],
        },
    ],
    "freezpak": [
        {"name": "Ancillary",           "key": "freezpak_ancillary",         "placeholder": False},
        {"name": "Inbound",             "key": "freezpak_inbound",           "placeholder": False},
        {"name": "Recurring / Storage", "key": "freezpak_recurring_storage", "placeholder": False},
        {"name": "XDock",               "key": "freezpak_xdock",             "placeholder": False},
    ],
    "exp": [
        {"name": "Inout Billing",        "key": "exp_inout_billing",      "placeholder": False},
        {"name": "Rejected & Repacking", "key": "exp_rejected_repacking", "placeholder": True},
    ],
    "efh": [
        {"name": "CGA",    "key": "efh_cga",    "placeholder": False},
        {"name": "Haines", "key": "efh_haines", "placeholder": False},
        {"name": "Streets","key": "efh_streets","placeholder": False},
        {"name": "Stults", "key": "efh_stults", "placeholder": False},
    ],
}

VALIDATED_INVOICE_KEYS = {"exp_inout_billing"}
EFH_INVOICE_KEYS           = {"efh_cga", "efh_haines", "efh_streets", "efh_stults"}
FREEZPAK_AGGREGATE_KEYS    = {"freezpak_ancillary", "freezpak_inbound", "freezpak_recurring_storage", "freezpak_xdock"}
HALLS_AGGREGATE_KEYS       = {
    "halls_trucking_fsc",
    "halls_warehousing_ancillary", "halls_warehousing_inbound",
    "halls_warehousing_renewal",   "halls_warehousing_sort_selection",
}
HALLS_INVOICE_KEYS         = {"halls_warehousing_ancillary_invoice", "freezpak_ancillary_invoice"}
FREEZPAK_INVOICE_PROC_TYPES = [
    {"name": "Ancillary", "key": "freezpak_ancillary_invoice", "placeholder": False},
]  # Add more Freezpak invoice types here as they are implemented


# =============================================================================
# SHARED HELPERS
# =============================================================================

def _norm(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).lower())

def _normalize_col(name: str) -> str:
    return re.sub(r"[\s_\-]+", " ", str(name).strip().lower())

def _find_column(ws, keyword: str):
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=1, column=col).value
        if raw is not None and keyword in _norm(raw):
            return col
    return None

def _po_to_int(value):
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    try:
        return int(float(str(value).strip().replace(",", "")))
    except (ValueError, TypeError):
        return None

def _sum_column(ws, col: int, data_start: int = 2) -> float:
    total = 0.0
    for row in range(data_start, ws.max_row + 1):
        v = ws.cell(row=row, column=col).value
        if v is not None and v != "":
            try:
                total += float(v)
            except (TypeError, ValueError):
                pass
    return total

def _last_data_row(ws, data_start: int = 2) -> int:
    last = data_start - 1
    for row in range(data_start, ws.max_row + 1):
        if any(ws.cell(row=row, column=c).value is not None for c in range(1, ws.max_column + 1)):
            last = row
    return last

def _read_first_sheet(uploaded_file):
    try:
        if uploaded_file.name.lower().endswith(".csv"):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file, sheet_name=0, dtype=str)
        uploaded_file.seek(0)
        return df, ""
    except Exception as exc:
        return None, str(exc)

def _read_all_sheets(uploaded_file):
    """Read every sheet from an Excel file (or the single sheet of a CSV).
    Returns (list_of_(sheet_name, df), error_string).
    """
    try:
        if uploaded_file.name.lower().endswith(".csv"):
            df = pd.read_csv(uploaded_file)
            uploaded_file.seek(0)
            return [("Sheet1", df)], ""
        sheets = pd.read_excel(uploaded_file, sheet_name=None, dtype=str)
        uploaded_file.seek(0)
        return list(sheets.items()), ""
    except Exception as exc:
        return [], str(exc)

def _build_canonical_columns(dfs):
    canon = {}
    for df in dfs:
        for col in df.columns:
            n = _normalize_col(col)
            if n not in canon:
                canon[n] = str(col).strip()
    return canon

def _align_df(df, canon):
    rename_map = {}
    for col in df.columns:
        n = _normalize_col(col)
        if n in canon:
            rename_map[col] = canon[n]
    return df.rename(columns=rename_map)

def _df_to_formatted_excel(df, sheet_name="Output"):
    """DataFrame → formatted Excel with numeric coercion, per-column widths, header style."""
    buf = io.BytesIO()
    wb  = Workbook()
    ws  = wb.active
    ws.title = sheet_name
    for ci, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=ci, value=str(col))
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=_coerce_numeric(val))
    _apply_output_formatting(ws)
    wb.save(buf)
    return buf.getvalue()

def to_excel_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Processed")
    return buf.getvalue()


# =============================================================================
# INVOICE PROCESSORS
# =============================================================================

def process_validated_invoice(uploaded_file) -> tuple:
    if not FREEZPAK_ENABLED:
        return None, None, "validate_invoice.py not found. Place it next to app.py and restart."
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path   = Path(tmpdir)
        input_path = tmp_path / uploaded_file.name
        input_path.write_bytes(uploaded_file.getbuffer())
        result_code = validate_run(input_path, tmp_path)
        if result_code == 2:
            exc_files = list(tmp_path.glob("*Exception Report*.xlsx"))
            if exc_files:
                return exc_files[0].read_bytes(), exc_files[0].name, "Rate validation failed -- Exception Report generated. Fix the flagged rows and resubmit."
            return None, None, "Rate validation failed but no exception report was generated."
        if result_code == 1:
            return None, None, "Processing failed -- check that your file has both a Detail tab and an AP tab."
        ap_files = list(tmp_path.glob("* - AP.xlsx"))
        if ap_files:
            return ap_files[0].read_bytes(), ap_files[0].name, None
        return None, None, "Processing completed but output file could not be found."


def process_efh_invoice(uploaded_file) -> tuple:
    file_bytes = uploaded_file.getbuffer()
    try:
        wb = load_workbook(io.BytesIO(bytes(file_bytes)), data_only=True)
    except Exception as exc:
        return None, None, f"Could not open workbook: {exc}"

    sheet_map    = {s.lower(): s for s in wb.sheetnames}
    ap_name      = sheet_map.get("ap", wb.sheetnames[0])
    details_name = sheet_map.get("details", sheet_map.get("detail", None))
    ws_ap        = wb[ap_name]

    ap_po_col      = _find_column(ws_ap, "po")
    ap_freight_col = _find_column(ws_ap, "freight")
    if ap_freight_col is None:
        return None, None, "Could not find a 'Freight Amount' column on the AP tab."

    if ap_po_col is not None:
        for row in range(2, ws_ap.max_row + 1):
            cell = ws_ap.cell(row=row, column=ap_po_col)
            if cell.value is not None:
                cell.value = _po_to_int(cell.value)

    if details_name is not None:
        ws_det = wb[details_name]
        det_freight_col = _find_column(ws_det, "freight")
        if det_freight_col is None:
            return None, None, "Could not find a 'Freight Amount' column on the Details tab."
        det_po_col = _find_column(ws_det, "po")
        if det_po_col is not None:
            for row in range(2, ws_det.max_row + 1):
                cell = ws_det.cell(row=row, column=det_po_col)
                if cell.value is not None:
                    cell.value = _po_to_int(cell.value)
        ap_total  = round(_sum_column(ws_ap,  ap_freight_col),  2)
        det_total = round(_sum_column(ws_det, det_freight_col), 2)
        if abs(ap_total - det_total) > 0.01:
            exc_wb = Workbook()
            exc_ws = exc_wb.active
            exc_ws.title = "Freight Mismatch"
            exc_ws.append(["Check", "Amount"])
            exc_ws.append(["AP Freight Total",      ap_total])
            exc_ws.append(["Details Freight Total", det_total])
            exc_ws.append(["Difference",            round(ap_total - det_total, 2)])
            buf = io.BytesIO()
            exc_wb.save(buf)
            stem = Path(uploaded_file.name).stem
            return buf.getvalue(), f"{stem} - Exception Report.xlsx", f"Freight totals do not match -- AP: ${ap_total:,.2f} vs Details: ${det_total:,.2f} (difference: ${abs(ap_total - det_total):,.2f}). Exception Report generated."

    last_row = _last_data_row(ws_ap)
    last_col = ws_ap.max_column
    center   = Alignment(horizontal="center", vertical="center", wrap_text=False)
    bold     = Font(bold=True)
    normal   = Font(bold=False)
    col_widths = []
    for col in range(1, last_col + 1):
        max_len = 0
        for row in range(1, last_row + 1):
            v = ws_ap.cell(row=row, column=col).value
            if v is not None:
                text_len = len(str(v).split(" ")[0]) if hasattr(v, "strftime") else len(str(v))
                max_len = max(max_len, text_len)
        col_widths.append(max_len)
    uniform_width = max(min(max(col_widths, default=10) + 2, 25), 10)
    for col in range(1, last_col + 1):
        ws_ap.column_dimensions[get_column_letter(col)].width = uniform_width
        for row in range(1, last_row + 1):
            cell = ws_ap.cell(row=row, column=col)
            cell.alignment = center
            cell.font = bold if row == 1 else normal
    ws_ap.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"

    if details_name is not None:
        del wb[details_name]
    for sheet in list(wb.sheetnames):
        if sheet != ap_name:
            del wb[sheet]

    buf = io.BytesIO()
    wb.save(buf)
    stem = Path(uploaded_file.name).stem
    return buf.getvalue(), f"{stem}-AP.xlsx", None


def process_generic(uploaded_file, xdock_label: str) -> tuple:
    try:
        df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith(".csv") else pd.read_excel(uploaded_file)
    except Exception as exc:
        return None, None, f"Could not read file: {exc}"
    df.columns = [str(c).strip().upper().replace(" ", "_") for c in df.columns]
    df = df.dropna(how="all")
    df["XDOCK_SOURCE"]   = xdock_label
    df["PROCESSED_DATE"] = datetime.now().strftime("%Y-%m-%d")
    stem = Path(uploaded_file.name).stem
    return to_excel_bytes(df), f"{stem} - AP.xlsx", None



# =============================================================================
# FREEZPAK AGGREGATE PROCESSORS
# =============================================================================

# Header style constants (shared by all output formatters)
_HDR_FILL = PatternFill("solid", fgColor="FF1E3A5F")  # dark navy (ARGB), matches dashboard theme
_HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_DAT_FONT = Font(bold=False, name="Calibri", size=10)
_CTR      = Alignment(horizontal="center", vertical="center", wrap_text=False)


def _coerce_numeric(val):
    """Try to convert a value to int or float; leave strings/None unchanged."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip().replace(",", "")
    if s == "" or s.lower() in ("nan", "none", "n/a", "-"):
        return None
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return val


def _apply_output_formatting(ws, sample_rows=200):
    """
    Per-column auto-fit widths, highlighted header row, auto-filter, freeze row 1.
    sample_rows: max rows scanned for width calculation (keeps large sheets fast).
    """
    last_row = ws.max_row
    last_col = ws.max_column
    scan_to  = min(last_row, sample_rows)

    for ci in range(1, last_col + 1):
        col_letter = get_column_letter(ci)
        # Per-column width: sample first sample_rows rows
        max_len = max(
            (len(str(ws.cell(row=r, column=ci).value or "")) for r in range(1, scan_to + 1)),
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)

        # Style every cell in this column
        for ri in range(1, last_row + 1):
            cell = ws.cell(row=ri, column=ci)
            if ri == 1:
                cell.fill      = _HDR_FILL
                cell.font      = _HDR_FONT
                cell.alignment = _CTR
            else:
                cell.font      = _DAT_FONT
                cell.alignment = _CTR

    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}1"
    ws.freeze_panes    = "A2"


def _freezpak_read_file(f, sheet_name):
    """Try named sheet first; fall back to first sheet."""
    try:
        df = pd.read_excel(f, sheet_name=sheet_name, dtype=str)
        f.seek(0)
        return df, ""
    except Exception:
        try:
            f.seek(0)
            df = pd.read_excel(f, sheet_name=0, dtype=str)
            f.seek(0)
            return df, ""
        except Exception as exc:
            return None, str(exc)


def _write_two_sheet_excel(df_raw, df_clean, sheet_raw="raw", sheet_clean="withoutSubtotals"):
    """Write two DataFrames to separate sheets with numeric coercion and formatting."""
    buf = io.BytesIO()
    wb  = Workbook()
    ws1 = wb.active
    ws1.title = sheet_raw
    for ci, col in enumerate(df_raw.columns, 1):
        ws1.cell(row=1, column=ci, value=str(col))
    for ri, row in enumerate(df_raw.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws1.cell(row=ri, column=ci, value=_coerce_numeric(val))
    _apply_output_formatting(ws1)

    ws2 = wb.create_sheet(title=sheet_clean)
    for ci, col in enumerate(df_clean.columns, 1):
        ws2.cell(row=1, column=ci, value=str(col))
    for ri, row in enumerate(df_clean.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws2.cell(row=ri, column=ci, value=_coerce_numeric(val))
    _apply_output_formatting(ws2)

    wb.save(buf)
    return buf.getvalue()


def process_freezpak_aggregate(uploaded_files, invoice_type_key: str) -> tuple:
    """
    Dispatcher for all three Freezpak aggregate types.
    Each handler returns (df_raw, df_clean, error_msg).
    """
    handlers = {
        "freezpak_ancillary":         _fpk_ancillary,
        "freezpak_inbound":           _fpk_inbound,
        "freezpak_recurring_storage": _fpk_storage,
        "freezpak_xdock":             _fpk_xdock,
    }
    handler = handlers.get(invoice_type_key)
    if not handler:
        return None, None, "No handler found for this invoice type."

    dfs_raw, dfs_clean, errors = [], [], []
    for f in uploaded_files:
        df_raw, df_clean, err = handler(f)
        if err:
            errors.append(f"{f.name}: {err}")
        else:
            dfs_raw.append(df_raw)
            dfs_clean.append(df_clean)

    if errors and not dfs_raw:
        return None, None, "\n".join(errors)

    df_raw_all   = pd.concat(dfs_raw,  ignore_index=True) if dfs_raw  else pd.DataFrame()
    df_clean_all = pd.concat(dfs_clean, ignore_index=True) if dfs_clean else pd.DataFrame()

    stem     = Path(uploaded_files[0].name).stem
    filename = f"{stem} - Aggregated.xlsx"
    out_bytes = _write_two_sheet_excel(df_raw_all, df_clean_all)
    msg = ("\n".join(errors)) if errors else None
    return out_bytes, filename, msg


def _fpk_xdock(f):
    """Freezpak XDock: Consolidated sheet, remove row 0, filter Trip totals, 2-sheet output."""
    df, err = _freezpak_read_file(f, "Consolidated")
    if err:
        return None, None, err
    df.columns = df.columns.str.strip()
    df = _rename_safe(df, {
        "Invoice": "InvoiceNum", "Invoice.1": "InvoiceDate",
        "Invoice.2": "InvoiceDueDate", "Total": "TotalPO",
        "Line": "LineHaul", "Fuel": "FuelSurcharge",
        "Detention": "DetentionCharge", "Stop": "StopCharge",
        "Cross": "Crossdock", "Pickup": "PickupCharge",
        "Toll": "TollCharge", "Storage": "StorageCharge",
        "Dock": "DockFee", "Adjusted": "AdjustedPallets"
    })
    if len(df) > 1:
        df = df.iloc[1:].reset_index(drop=True)
    df_raw = df.copy()
    if "Trip" in df.columns:
        df_clean = df[~df["Trip"].astype(str).str.contains("Total", na=False)].copy()
    else:
        df_clean = df.copy()
    col_order = ["Store", "City", "InvoiceNum", "InvoiceDate", "InvoiceDueDate", "PO",
                 "TotalPO", "LineHaul", "FuelSurcharge", "DetentionCharge", "StopCharge",
                 "Crossdock", "PickupCharge", "TollCharge", "StorageCharge", "DockFee",
                 "Quantity", "Weight", "Cube", "AdjustedPallets", "Buyer", "Trip"]
    existing = [c for c in col_order if c in df_clean.columns]
    extra    = [c for c in df_clean.columns if c not in existing]
    df_clean = df_clean[existing + extra]
    return df_raw, df_clean, None


def _rename_safe(df, rename_dict):
    existing = {k: v for k, v in rename_dict.items() if k in df.columns}
    return df.rename(columns=existing)


def _fpk_ancillary(f):
    """Freezpak Ancillary: Sheet1, remove row 0, filter PONum, recalc TotalPerPO."""
    df, err = _freezpak_read_file(f, "Sheet1")
    if err:
        return None, None, err
    df.columns = df.columns.str.strip()
    df = _rename_safe(df, {
        "Invoice": "InvoiceNum", "Invoice.1": "InvoiceDueDate",
        "PO": "FullPO", "PO.1": "PONum", "Invoice.2": "InvoiceAmount",
        "BOL": "BOLPreparation", "CROSS DOCK": "CrossDockOut",
        "CASE": "CaseSelection", "FLOOR": "FloorUnloading",
        "CATCH": "CatchWeight", "Total Per": "TotalPerPO"
    })
    df = df.iloc[1:].reset_index(drop=True)
    df_raw = df.copy()
    po_col = "PONum" if "PONum" in df.columns else (df.columns[0] if len(df.columns) else None)
    if po_col:
        df_clean = df[~df[po_col].astype(str).str.contains("Total", na=False)].copy()
        if len(df_clean) > 0:
            df_clean = df_clean.iloc[:-1]
    else:
        df_clean = df.copy()
    charge_cols = ["BOLPreparation", "CrossDockOut", "CaseSelection", "FloorUnloading", "CatchWeight"]
    existing = [c for c in charge_cols if c in df_clean.columns]
    if existing:
        df_clean[existing] = df_clean[existing].apply(pd.to_numeric, errors="coerce")
        df_clean["TotalPerPO"] = df_clean[existing].sum(axis=1)
    return df_raw, df_clean, None


def _fpk_inbound(f):
    """Freezpak Inbound: Sheet1, remove row 0, filter PO, recalc CrossTotal."""
    df, err = _freezpak_read_file(f, "Sheet1")
    if err:
        return None, None, err
    df.columns = df.columns.str.strip()
    df = _rename_safe(df, {
        "Invoice": "InvoiceNum", "Invoice.1": "InvoiceDate",
        "Invoice.2": "InvoiceDueDate", "Invoice.3": "InvoiceAmount",
        "Monthly": "MonthlyStorageperPL", "14 Day": "14DayRate",
        "Handling": "Handling(in and out per pallet)", "Cross": "CrossTotal"
    })
    df = df.iloc[1:].reset_index(drop=True)
    df_raw = df.copy()
    if "PO" in df.columns:
        df_clean = df[~df["PO"].astype(str).str.contains("Total", na=False)].copy()
    else:
        df_clean = df.copy()
    rate_cols = ["MonthlyStorageperPL", "14DayRate", "Handling(in and out per pallet)", "Floor unloading", "Catch Wgt"]
    existing = [c for c in rate_cols if c in df_clean.columns]
    if existing:
        df_clean[existing] = df_clean[existing].apply(pd.to_numeric, errors="coerce")
        df_clean["CrossTotal"] = df_clean[existing].sum(axis=1)
    col_order = ["Vendor", "InvoiceNum", "InvoiceDate", "InvoiceDueDate", "PO",
                 "InvoiceAmount", "Product", "Item", "Case", "Pallet", "Weight",
                 "MonthlyStorageperPL", "14DayRate", "Handling(in and out per pallet)",
                 "Floor unloading", "Catch Wgt", "CrossTotal"]
    existing_order = [c for c in col_order if c in df_clean.columns]
    extra = [c for c in df_clean.columns if c not in existing_order]
    df_clean = df_clean[existing_order + extra]
    return df_raw, df_clean, None


def _fpk_storage(f):
    """Freezpak Recurring/Storage: first sheet, remove row 0, filter PONum."""
    df, err = _freezpak_read_file(f, 0)
    if err:
        return None, None, err
    df.columns = df.columns.str.strip()
    df = _rename_safe(df, {
        "Invoice": "InvoiceNum", "Invoice.1": "InvoiceDate",
        "Invoice.2": "InvoiceDueDate", "Lot": "PONum",
        "Item": "ItemCode", "Item.1": "ItemDesc",
        "Invoice.3": "InvoiceAmount", "Initial": "InitialDate",
        "Recurring": "RecurringStorage", "Date": "DateFrom", "Date.1": "DateTo"
    })
    df = df.iloc[1:].reset_index(drop=True)
    df_raw = df.copy()
    if "PONum" in df.columns:
        df_clean = df[~df["PONum"].astype(str).str.contains("Total", na=False)].copy()
    else:
        df_clean = df.copy()
    col_order = ["Vendor", "InvoiceNum", "InvoiceDate", "InvoiceDueDate", "PO", "PONum",
                 "InvoiceAmount", "ItemCode", "ItemDesc", "Pallet",
                 "InitialDate", "RecurringStorage", "DateFrom", "DateTo"]
    existing_order = [c for c in col_order if c in df_clean.columns]
    extra = [c for c in df_clean.columns if c not in existing_order]
    df_clean = df_clean[existing_order + extra]
    return df_raw, df_clean, None


# =============================================================================
# HALLS AGGREGATE PROCESSORS
# =============================================================================

def _write_single_sheet_excel(df, sheet_name="Aggregated"):
    """Single-sheet formatted Excel output with numeric coercion."""
    buf = io.BytesIO()
    wb  = Workbook()
    ws  = wb.active
    ws.title = sheet_name
    for ci, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=ci, value=str(col))
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=_coerce_numeric(val))
    _apply_output_formatting(ws)
    wb.save(buf)
    return buf.getvalue()


def _halls_read_logistics(f):
    """Read a Halls LOGISTICS data file: drop Unnamed columns, blank rows, and last (totals) row."""
    try:
        df = pd.read_excel(f, sheet_name=0, dtype=str)
        f.seek(0)
        df = df.loc[:, ~df.columns.str.contains('^Unnamed', na=False)]
        df = df.dropna(how='all').reset_index(drop=True)   # remove blank rows
        if len(df) > 0:
            df = df.iloc[:-1]                               # remove last totals row
        return df, ""
    except Exception as exc:
        return None, str(exc)


def _halls_dedup_columns(dfs):
    """Collect all column names across DataFrames, deduplicated, order preserved."""
    seen, ordered = set(), []
    for df in dfs:
        for col in df.columns:
            if col not in seen:
                seen.add(col)
                ordered.append(col)
    return ordered


def _halls_simple_aggregate(uploaded_files):
    """
    Shared logic for Ancillary, Inbound, Sort&Selection, Renewal.
    LOGISTICS files are data; anything else is an AP summary (shown but not aggregated).
    Returns (merged_df, ap_info_list, errors).
    """
    dfs, ap_info, errors = [], [], []
    for f in uploaded_files:
        if "LOGISTICS" in f.name.upper():
            df, err = _halls_read_logistics(f)
            if err:
                errors.append(f"{f.name}: {err}")
            else:
                dfs.append(df)
        else:
            # AP/summary file -- read invoice header info only
            try:
                df_ap = pd.read_excel(f, sheet_name=0)
                f.seek(0)
                if len(df_ap) > 1:
                    row = df_ap.iloc[1]
                    inv  = next((row.get(k, None) for k in ["Invoice#", "INVOICE", "InvoiceNum"] if k in df_ap.columns), "N/A")
                    date = next((row.get(k, None) for k in ["Invoice Date", "DATE", "InvoiceDate"] if k in df_ap.columns), "N/A")
                    due  = next((row.get(k, None) for k in ["Due Date", "DUE DATE", "InvoiceDueDate"] if k in df_ap.columns), "N/A")
                    tot  = round(df_ap["Total"].sum(), 2) if "Total" in df_ap.columns else 0
                    ap_info.append({"File": f.name, "Invoice": inv, "Invoice Date": date, "Due Date": due, "Total": f"${tot:,.2f}"})
            except Exception as exc:
                errors.append(f"{f.name} (AP read): {exc}")
    return dfs, ap_info, errors


def _halls_ancillary_proc(files):
    dfs, ap_info, errors = _halls_simple_aggregate(files)
    if not dfs:
        return None, None, "\n".join(errors) or "No LOGISTICS files found. Files with LOGISTICS in the name are used as data sources."
    col_order = _halls_dedup_columns(dfs)
    merged = pd.concat(dfs, ignore_index=True)
    existing = [c for c in col_order if c in merged.columns]
    out = _write_single_sheet_excel(merged[existing])
    return out, "Halls_Ancillary_Aggregated.xlsx", "\n".join(errors) if errors else None


def _halls_inbound_proc(files):
    dfs, ap_info, errors = _halls_simple_aggregate(files)
    if not dfs:
        return None, None, "\n".join(errors) or "No LOGISTICS files found."
    col_order = _halls_dedup_columns(dfs)
    merged = pd.concat(dfs, ignore_index=True)
    existing = [c for c in col_order if c in merged.columns]
    out = _write_single_sheet_excel(merged[existing], "Inbound")
    return out, "Halls_Inbound_Aggregated.xlsx", "\n".join(errors) if errors else None


def _halls_sort_selection_proc(files):
    dfs, ap_info, errors = _halls_simple_aggregate(files)
    if not dfs:
        return None, None, "\n".join(errors) or "No LOGISTICS files found."
    col_order = _halls_dedup_columns(dfs)
    merged = pd.concat(dfs, ignore_index=True)
    existing = [c for c in col_order if c in merged.columns]
    out = _write_single_sheet_excel(merged[existing], "Sort&Selection")
    return out, "Halls_SortSelection_Aggregated.xlsx", "\n".join(errors) if errors else None


def _halls_renewal_proc(files):
    dfs, ap_info, errors = _halls_simple_aggregate(files)
    if not dfs:
        return None, None, "\n".join(errors) or "No LOGISTICS files found."
    col_order = _halls_dedup_columns(dfs)
    merged = pd.concat(dfs, ignore_index=True)
    existing = [c for c in col_order if c in merged.columns]
    out = _write_single_sheet_excel(merged[existing], "Renewal")
    return out, "Halls_Renewal_Aggregated.xlsx", "\n".join(errors) if errors else None


def _halls_trucking_fsc_proc(files):
    """
    Halls Trucking & FSC (XDock notebook logic).
    Non-XDOCK files: read 'Logistics' sheet.
    Splits into 4 sheets: raw, withoutSubtotals, CPU (TL null), XDOCK (TL not null).
    """
    dfs_raw, errors = [], []
    for f in files:
        if "XDOCK" in f.name.upper():
            # AP/summary file for this type -- skip aggregation
            continue
        try:
            df = pd.read_excel(f, sheet_name="Logistics")
            f.seek(0)
            dfs_raw.append(df)
        except Exception:
            try:
                f.seek(0)
                df = pd.read_excel(f, sheet_name=0, dtype=str)
                f.seek(0)
                dfs_raw.append(df)
            except Exception as exc:
                errors.append(f"{f.name}: {exc}")

    if not dfs_raw:
        return None, None, "\n".join(errors) or "No Logistics sheet found in uploaded files."

    col_order = _halls_dedup_columns(dfs_raw)
    df_raw   = pd.concat(dfs_raw, ignore_index=True)
    existing = [c for c in col_order if c in df_raw.columns]
    df_raw   = df_raw[existing]

    if "TL" in df_raw.columns:
        tl = df_raw["TL"].astype(str)
        df_clean = df_raw[~tl.str.contains("Total", na=False) & ~tl.str.contains("^TL$", na=False)]
        df_cpu   = df_clean[df_clean["TL"].isna() | (df_clean["TL"].astype(str).str.strip() == "nan")]
        df_xdock = df_clean[df_clean["TL"].notna() & (df_clean["TL"].astype(str).str.strip() != "nan")]
    else:
        df_clean = df_raw.copy()
        df_cpu   = pd.DataFrame(columns=df_raw.columns)
        df_xdock = pd.DataFrame(columns=df_raw.columns)

    buf = io.BytesIO()
    wb  = Workbook()

    for sheet_name, df in [("raw", df_raw), ("withoutSubtotals", df_clean),
                            ("CPU", df_cpu), ("XDOCK", df_xdock)]:
        ws = wb.active if sheet_name == "raw" else wb.create_sheet(title=sheet_name)
        ws.title = sheet_name
        for ci, col in enumerate(df.columns, 1):
            ws.cell(row=1, column=ci, value=str(col))
        for ri, row in enumerate(df.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=_coerce_numeric(val))
        _apply_output_formatting(ws)

    wb.save(buf)
    return buf.getvalue(), "Halls_TruckingFSC_Aggregated.xlsx", "\n".join(errors) if errors else None


def process_halls_ancillary_invoice(uploaded_file) -> tuple:
    """
    Process the Halls Warehousing Ancillary Invoice template (single .xlsx file).

    Sheet1:
      - Validate/fix BOL (col M) on Total rows: 5.94 auto-divided to 2.97; >2.97 flagged.
      - Compute Case Selection (col T) = O/J; >0.240 flagged.
      - Write computed values into col S (BOL) and col T (Case Selection).

    Invoice tab:
      - Compute STORE (col H) = first 3 digits of PO (col E), written as values.
      - Collect unique stores in appearance order.
      - Get base invoice number from B2.
      - Assign suffix: stores 1-45 → InvoiceNum-1, 46-90 → InvoiceNum-2, etc.
      - Rebuild Sheet2 with STORE / Invoice# mapping.
      - Replace col B with suffixed invoice numbers.
    """
    f = uploaded_file[0] if isinstance(uploaded_file, list) else uploaded_file
    try:
        raw_bytes = f.read()
        f.seek(0)
        wb = load_workbook(io.BytesIO(raw_bytes))
    except Exception as exc:
        return None, None, f"Could not open workbook: {exc}"

    # ------------------------------------------------------------------ Sheet1
    if "Sheet1" not in wb.sheetnames:
        return None, None, "Expected a 'Sheet1' tab — not found in this file."
    ws1 = wb["Sheet1"]

    bol_flags  = []  # (row_num, original_value)  — needs manual review
    cs_flags   = []  # (row_num, computed_value)   — exceeds 0.240
    bol_fixed  = []  # rows where 5.94 was auto-corrected to 2.97

    # Header rows are 1 and 2; data starts at row 3.
    # Column index mapping (0-based within row tuple):
    #   F=5, J=9, M=12, O=14, S=18, T=19
    for row in ws1.iter_rows(min_row=3, max_row=ws1.max_row):
        f_cell = row[5]   # PO# / "XXXXX Total"
        m_cell = row[12]  # BOL Preparation amount
        o_cell = row[14]  # Case Selection amount
        j_cell = row[9]   # Cases count
        s_cell = row[18]  # BOL computed output
        t_cell = row[19]  # Case Selection computed output

        f_str      = str(f_cell.value).strip() if f_cell.value is not None else ""
        f_upper    = f_str.upper()
        # PO-level Total rows look like "70223 Total"; skip "Grand Total" catch-all row
        is_total   = f_upper.endswith("TOTAL") and f_upper != "GRAND TOTAL"

        # --- BOL: only on PO Total rows ---
        if is_total and m_cell.value is not None:
            try:
                bol     = float(m_cell.value)
                bol_r   = round(bol, 4)          # eliminate float noise (e.g. 2.9700000000000006)
                if round(bol, 2) == 5.94:
                    # Forgot to split into 2 cases — auto-correct to 2.97
                    m_cell.value = 2.97
                    s_cell.value = 2.97
                    bol_fixed.append(f_cell.row)
                elif bol_r > 2.97:
                    s_cell.value = bol_r
                    bol_flags.append((f_cell.row, bol_r))
                else:
                    s_cell.value = bol_r
            except (TypeError, ValueError):
                s_cell.value = ""
        elif not is_total:
            s_cell.value = ""

        # --- Case Selection: all data rows ---
        try:
            o_val = float(o_cell.value) if o_cell.value is not None else None
            j_val = float(j_cell.value) if j_cell.value is not None else None
            if o_val is not None and j_val and j_val != 0:
                cs = round(o_val / j_val, 6)
                t_cell.value = cs
                if cs > 0.240:
                    cs_flags.append((f_cell.row, round(cs, 4)))
            else:
                t_cell.value = ""
        except (TypeError, ValueError):
            t_cell.value = ""

    # ---------------------------------------------------------------- Invoice tab
    if "Invoice" not in wb.sheetnames:
        return None, None, "Expected an 'Invoice' tab — not found in this file."
    wsi = wb["Invoice"]

    # Base invoice number (cell B2, first data row)
    raw_inv = wsi.cell(2, 2).value
    try:
        base_inv = str(int(float(raw_inv)))
    except (TypeError, ValueError):
        base_inv = str(raw_inv).strip() if raw_inv else "UNKNOWN"

    # Walk Invoice rows: compute STORE, collect unique stores in appearance order
    store_list  = []          # ordered unique stores
    seen_stores = set()
    row_store   = {}          # row_index → store string

    for r in range(2, wsi.max_row + 1):
        e_val = wsi.cell(r, 5).value   # PO number
        if e_val is None:
            continue
        try:
            po_str = str(int(float(e_val)))
        except (TypeError, ValueError):
            po_str = str(e_val).strip()
        store = po_str[:3]
        wsi.cell(r, 8).value = store   # write STORE as a plain value
        row_store[r] = store
        if store not in seen_stores:
            store_list.append(store)
            seen_stores.add(store)

    # Assign suffixed invoice numbers (groups of 45 stores)
    store_to_inv = {}
    for idx, store in enumerate(store_list):
        suffix = (idx // 45) + 1
        store_to_inv[store] = f"{base_inv}-{suffix}"

    # Replace Invoice col B with suffixed invoice numbers
    for r, store in row_store.items():
        wsi.cell(r, 2).value = store_to_inv[store]

    # ---------------------------------------------------------------- Cleanup helpers
    # Remove STORE helper column (col H) from Invoice tab — no longer needed
    wsi.delete_cols(8)

    # Remove Sheet2 helper tab from workbook
    if "Sheet2" in wb.sheetnames:
        wb.remove(wb["Sheet2"])

    # ---------------------------------------------------------------- Format output sheets
    # Preserve original template styling — only add AutoFilter, freeze panes,
    # and auto-fit column widths (no header colour/font overrides).
    for _sn in wb.sheetnames:
        _ws = wb[_sn]
        _lc = _ws.max_column
        _lr = _ws.max_row
        scan_to = min(_lr, 200)
        # Auto-fit column widths
        for _ci in range(1, _lc + 1):
            _col_letter = get_column_letter(_ci)
            _max_len = max(
                (len(str(_ws.cell(row=_r, column=_ci).value or "")) for _r in range(1, scan_to + 1)),
                default=8,
            )
            _ws.column_dimensions[_col_letter].width = min(max(_max_len + 2, 8), 40)
        # AutoFilter on header row
        if _sn == "Sheet1":
            _ws.auto_filter.ref = f"A1:{get_column_letter(_lc)}1"
            _ws.freeze_panes = "A3"   # Sheet1 has 2 header rows
        else:
            _ws.auto_filter.ref = f"A1:{get_column_letter(_lc)}1"
            _ws.freeze_panes = "A2"

    # ---------------------------------------------------------------- Save
    buf = io.BytesIO()
    wb.save(buf)

    # Build human-readable summary
    n_groups = ((len(store_list) - 1) // 45 + 1) if store_list else 0
    lines = [
        f"Base invoice: {base_inv}",
        f"Unique stores: {len(store_list)}  →  {n_groups} invoice group(s)",
    ]
    if bol_fixed:
        lines.append(f"✓ BOL auto-corrected (5.94 → 2.97) on {len(bol_fixed)} row(s): rows {bol_fixed}")
    if bol_flags:
        lines.append(
            f"⚠️ BOL exceeds $2.97 — manual review required: "
            + ", ".join(f"row {r} = ${v}" for r, v in bol_flags)
        )
    if cs_flags:
        lines.append(
            f"⚠️ Case Selection exceeds 0.240 — manual review required: "
            + ", ".join(f"row {r} = {v}" for r, v in cs_flags)
        )
    if not bol_flags and not cs_flags:
        lines.append("✓ All BOL and Case Selection values are within limits")

    return buf.getvalue(), f"Ancillary_Invoice_{base_inv}.xlsx", "\n".join(lines)


def process_halls_aggregate(uploaded_files, invoice_type_key: str) -> tuple:
    handlers = {
        "halls_warehousing_ancillary":      _halls_ancillary_proc,
        "halls_warehousing_inbound":        _halls_inbound_proc,
        "halls_warehousing_sort_selection": _halls_sort_selection_proc,
        "halls_warehousing_renewal":        _halls_renewal_proc,
        "halls_trucking_fsc":               _halls_trucking_fsc_proc,
    }
    handler = handlers.get(invoice_type_key)
    if not handler:
        return None, None, f"No handler configured for {invoice_type_key}."
    return handler(uploaded_files)


def route_invoice(uploaded_files, xdock_key: str, invoice_type_key: str) -> tuple:
    # uploaded_files may be a single UploadedFile or a list
    if invoice_type_key in VALIDATED_INVOICE_KEYS:
        f = uploaded_files if not isinstance(uploaded_files, list) else uploaded_files[0]
        return process_validated_invoice(f)
    if invoice_type_key in EFH_INVOICE_KEYS:
        f = uploaded_files if not isinstance(uploaded_files, list) else uploaded_files[0]
        return process_efh_invoice(f)
    if invoice_type_key in FREEZPAK_AGGREGATE_KEYS:
        files = uploaded_files if isinstance(uploaded_files, list) else [uploaded_files]
        return process_freezpak_aggregate(files, invoice_type_key)
    if invoice_type_key in HALLS_AGGREGATE_KEYS:
        files = uploaded_files if isinstance(uploaded_files, list) else [uploaded_files]
        return process_halls_aggregate(files, invoice_type_key)
    if invoice_type_key in HALLS_INVOICE_KEYS:
        f = uploaded_files if not isinstance(uploaded_files, list) else uploaded_files[0]
        return process_halls_ancillary_invoice(f)
    return None, None, "PLACEHOLDER"


# =============================================================================
# INVOICE SECTION RENDERER
# =============================================================================

def render_invoice_section(xdock_key, invoice_type_cfg, xdock_color, xdock_display):
    inv_key      = invoice_type_cfg["key"]
    inv_name     = invoice_type_cfg["name"]
    is_ph        = invoice_type_cfg["placeholder"]
    is_validated = (inv_key in VALIDATED_INVOICE_KEYS)
    is_efh       = (inv_key in EFH_INVOICE_KEYS)
    is_fpk_agg        = (inv_key in FREEZPAK_AGGREGATE_KEYS)
    is_halls_agg      = (inv_key in HALLS_AGGREGATE_KEYS)
    is_ancillary_inv  = (inv_key in HALLS_INVOICE_KEYS)
    is_multi          = is_fpk_agg or is_halls_agg

    if is_ph:
        st.markdown('<div class="placeholder-box"><b>Logic not configured yet.</b> Upload and processing workflow will be added later. You can still upload a file to preview its contents.</div>', unsafe_allow_html=True)
    elif is_validated:
        if not FREEZPAK_ENABLED:
            st.markdown('<div class="warning-box"><b>validate_invoice.py not detected.</b> Place it next to app.py and restart.</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="info-box"><b>How to use:</b><br><span class="step-badge">1</span>Upload your {xdock_display} &mdash; {inv_name} workbook (.xlsx) &mdash; must have <b>Detail</b> and <b>AP</b> tabs<br><span class="step-badge">2</span>Review the raw data preview<br><span class="step-badge">3</span>Click <b>Process Invoice</b> &mdash; rate check runs automatically (max $0.475/lb)<br><span class="step-badge">4</span>Download the <b>AP-only output</b>, or the <b>Exception Report</b> if rates fail</div>', unsafe_allow_html=True)
    elif is_efh:
        st.markdown(f'<div class="info-box"><b>How to use:</b><br><span class="step-badge">1</span>Upload your {inv_name} invoice workbook (.xlsx) &mdash; should have an <b>AP</b> tab and optionally a <b>Details</b> tab<br><span class="step-badge">2</span>Review the raw data preview<br><span class="step-badge">3</span>Click <b>Process Invoice</b> &mdash; PO numbers normalised, freight totals validated<br><span class="step-badge">4</span>Download the formatted <b>AP-only output</b>, or the <b>Exception Report</b> if totals do not match</div>', unsafe_allow_html=True)
    elif is_ancillary_inv:
        # ── How to use (full width) ──────────────────────────────────────
        st.markdown(
            '<div class="info-box">'
            '<b>How to use</b><br>'
            '<span class="step-badge">1</span>Upload your Ancillary Invoice template (.xlsx) &mdash; '
            'must have <b>Sheet1</b> and <b>Invoice</b> tabs<br>'
            '<span class="step-badge">2</span>Click <b>Process Invoice</b> to run all checks and transformations automatically<br>'
            '<span class="step-badge">3</span>Download the processed workbook and review the summary for any flagged rows'
            '</div>',
            unsafe_allow_html=True,
        )
        # ── Validations + Transformations side by side (pure HTML flexbox) ───
        st.markdown(
            '<div style="display:flex;gap:12px;margin-top:8px">'
            '<div class="warning-box" style="flex:1;margin:0">'
            '<b>&#9888;&nbsp; Validation Checks &mdash; Sheet1</b><br><br>'
            '<b>BOL</b> &nbsp;&mdash;&nbsp; Column M on Total rows<br>'
            '&nbsp;&nbsp;&bull;&nbsp; Max: <b>$2.97</b><br>'
            '&nbsp;&nbsp;&bull;&nbsp; If <b>$5.94</b> found: auto-corrected to $2.97 (forgotten case split)<br>'
            '&nbsp;&nbsp;&bull;&nbsp; Any other value above $2.97: <b>flagged for review</b><br><br>'
            '<b>Case Selection</b> &nbsp;&mdash;&nbsp; Case Selection Amount &divide; Case<br>'
            '&nbsp;&nbsp;&bull;&nbsp; Max: <b>0.240</b><br>'
            '&nbsp;&nbsp;&bull;&nbsp; Values above 0.240: <b>flagged for review</b><br>'
            '&nbsp;&nbsp;&bull;&nbsp; Division error (blank Cases): left blank'
            '</div>'
            '<div class="success-box" style="flex:1;margin:0">'
            '<b>&#10003;&nbsp; Transformations Applied &mdash; Invoice tab</b><br><br>'
            '<b>STORE column</b><br>'
            '&nbsp;&nbsp;&bull;&nbsp; First 3 digits of PO number written as plain values<br><br>'
            '<b>Invoice suffix assignment</b> &nbsp;&mdash;&nbsp; Column B<br>'
            '&nbsp;&nbsp;&bull;&nbsp; Unique stores collected in order of appearance<br>'
            '&nbsp;&nbsp;&bull;&nbsp; Batched in groups of 45: stores 1&ndash;45 &rarr; <b>InvoiceNum&#8209;1</b>, '
            '46&ndash;90 &rarr; <b>InvoiceNum&#8209;2</b>, and so on<br>'
            '&nbsp;&nbsp;&bull;&nbsp; Column B updated with suffixed invoice numbers<br>'
            '&nbsp;&nbsp;&bull;&nbsp; Helper STORE column and Sheet2 <b>removed from output</b>'
            '</div>'
            '</div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(f'<div class="info-box"><b>How to use:</b><br><span class="step-badge">1</span>Upload your {xdock_display} &mdash; {inv_name} invoice file (Excel or CSV)<br><span class="step-badge">2</span>Review the raw data preview<br><span class="step-badge">3</span>Click <b>Process Invoice</b> to run the automation<br><span class="step-badge">4</span>Download the output file</div>', unsafe_allow_html=True)

    accepted_types = ["xlsx", "xls"] if (is_validated or is_efh or is_fpk_agg or is_halls_agg or is_ancillary_inv) else ["xlsx", "xls", "csv"]
    file_hint = ("Excel (.xlsx) with AP tab (+ optional Details tab)" if is_efh
                 else "Excel (.xlsx) with Detail and AP tabs" if is_validated
                 else "Ancillary Invoice template (.xlsx) — must have Sheet1 and Invoice tabs" if is_ancillary_inv
                 else "Upload one or more Excel files (.xlsx)" if (is_fpk_agg or is_halls_agg)
                 else "Excel (.xlsx, .xls) or CSV")

    uploaded = st.file_uploader(
        label=f"Upload {xdock_display} -- {inv_name}",
        type=accepted_types,
        key=f"upload_{inv_key}",
        label_visibility="collapsed",
        accept_multiple_files=is_multi,
    )

    # Normalise to list for uniform handling
    uploaded_list = uploaded if isinstance(uploaded, list) else ([uploaded] if uploaded else [])
    if uploaded_list:
        uploaded = uploaded_list  # always a list from here
        try:
            # Read only 100 rows for preview to keep memory low
            raw_df = (pd.read_csv(uploaded[0], nrows=100) if uploaded[0].name.endswith(".csv")
                      else pd.read_excel(uploaded[0], sheet_name=0, nrows=100))
            uploaded[0].seek(0)
        except Exception as exc:
            st.markdown(f'<div class="warning-box">Could not preview file: {exc}</div>', unsafe_allow_html=True)
            return

        sheet_label = "Detail tab rows" if is_validated else "AP tab rows" if is_efh else "Rows (file 1)" if (is_fpk_agg or is_halls_agg) else "Total Rows"

        freight_total_html = ""
        for col in raw_df.columns:
            col_norm = re.sub(r"[^a-z0-9]", "", str(col).lower())
            if "freight" in col_norm or "amount" in col_norm:
                try:
                    total = pd.to_numeric(raw_df[col], errors="coerce").sum()
                    freight_total_html = (
                        '<div class="metric-card" style="border-color:#4fc3f7;background:#132030">'
                        '<div class="metric-value" style="color:#4fc3f7">$' + f'{total:,.2f}' + '</div>'
                        '<div class="metric-label">' + str(col) + ' Total</div>'
                        '</div>'
                    )
                    break
                except Exception:
                    pass

        st.markdown(f"""
        <div class="metric-row">
            <div class="metric-card"><div class="metric-value">{len(uploaded) if isinstance(uploaded, list) else 1}</div><div class="metric-label">{"Files" if isinstance(uploaded, list) and len(uploaded) > 1 else sheet_label}</div></div>
            <div class="metric-card"><div class="metric-value">{len(raw_df):,}</div><div class="metric-label">{sheet_label}</div></div>
            <div class="metric-card"><div class="metric-value">{len(raw_df.columns)}</div><div class="metric-label">Columns</div></div>
            <div class="metric-card"><div class="metric-value">{raw_df.isnull().sum().sum()}</div><div class="metric-label">Blank Cells</div></div>
            <div class="metric-card"><div class="metric-value">{uploaded[0].name.split(".")[-1].upper() if isinstance(uploaded, list) else uploaded.name.split(".")[-1].upper()}</div><div class="metric-label">File Type</div></div>{freight_total_html}
        </div>
        """, unsafe_allow_html=True)

        with st.expander("\U0001f4cb Raw Data Preview (first sheet)", expanded=True):
            st.dataframe(raw_df.head(20), use_container_width=True, height=280)

        st.markdown("<br>", unsafe_allow_html=True)

        if not is_ph:
            sess_processed = f"processed_{inv_key}"
            sess_output    = f"output_{inv_key}"

            col1, _col2, _col3 = st.columns([1, 1, 2])
            with col1:
                process_clicked = st.button("Process Invoice", key=f"process_{inv_key}")

            if process_clicked:
                st.session_state.pop(sess_output, None)
                st.session_state[sess_processed] = True

            if st.session_state.get(sess_processed):
                if sess_output not in st.session_state:
                    with st.spinner("Running validation and processing..."):
                        for uf in uploaded:
                            uf.seek(0)
                        st.session_state[sess_output] = route_invoice(uploaded, xdock_key, inv_key)

                output_bytes, output_filename, error_msg = st.session_state[sess_output]

                if error_msg and error_msg != "PLACEHOLDER":
                    is_warning = any(w in error_msg.lower() for w in ("failed", "exception", "mismatch", "do not match"))
                    box = "warning-box" if is_warning else "error-box"
                    st.markdown(f'<div class="{box}">{error_msg}</div>', unsafe_allow_html=True)

                if output_bytes:
                    if not error_msg:
                        st.markdown('<div class="success-box">Validation passed -- output ready for download</div>', unsafe_allow_html=True)
                    col_a, _col_b, _col_c = st.columns([1, 1, 2])
                    with col_a:
                        st.download_button(
                            label="Download Output",
                            data=output_bytes,
                            file_name=output_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"download_{inv_key}",
                        )
    else:
        st.markdown(f"""
        <div style="text-align:center;padding:48px 0;color:#3a4a6a;border:1.5px dashed #2a3a5c;border-radius:12px;margin-top:12px;">
            <div style="font-size:40px;margin-bottom:12px">&#x2191;</div>
            <div style="font-family:\'IBM Plex Mono\',monospace;font-size:14px">Drop your {inv_name} invoice file here</div>
            <div style="font-size:12px;margin-top:6px">{file_hint}</div>
        </div>
        """, unsafe_allow_html=True)


# =============================================================================
# FILE AGGREGATOR TAB
# =============================================================================

def render_aggregator_tab():
    st.markdown("""
    <div style="margin-bottom:20px">
        <span style="font-family:\'IBM Plex Mono\',monospace;font-size:18px;font-weight:600;color:#b39ddb">\U0001f4c1  File Aggregator</span>
        <span style="font-size:12px;color:#7a8aaa;margin-left:12px">Combine multiple Excel files into one unified output</span>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="info-box"><b>How to use:</b><br>
    <span class="step-badge">1</span>Upload two or more Excel files &mdash; column order and extra columns are handled automatically<br>
    <span class="step-badge">2</span>Review the file summary<br>
    <span class="step-badge">3</span>Click <b>Aggregate Files</b> &mdash; rows are appended sequentially, columns aligned<br>
    <span class="step-badge">4</span>Download the single merged Excel file
    </div>
    """, unsafe_allow_html=True)

    uploaded_files = st.file_uploader(
        "Upload Excel files to aggregate",
        type=["xlsx", "xls", "csv"],
        accept_multiple_files=True,
        key="agg_uploader",
        label_visibility="collapsed",
    )

    if not uploaded_files:
        st.markdown("""
        <div style="text-align:center;padding:48px 0;color:#3a4a6a;border:1.5px dashed #2a3a5c;border-radius:12px;margin-top:12px;">
            <div style="font-size:40px;margin-bottom:12px">&#x2191;</div>
            <div style="font-family:\'IBM Plex Mono\',monospace;font-size:14px">Drop two or more files here to aggregate</div>
            <div style="font-size:12px;margin-top:6px">Excel (.xlsx, .xls) or CSV</div>
        </div>
        """, unsafe_allow_html=True)
        return

    dfs, file_meta, errors = [], [], []
    for f in uploaded_files:
        sheet_list, err = _read_all_sheets(f)
        if err:
            errors.append(f"<b>{f.name}</b>: {err}")
        else:
            for sheet_name, df in sheet_list:
                dfs.append(df)
                label = f.name if len(sheet_list) == 1 else f"{f.name}  [{sheet_name}]"
                file_meta.append({"File": label, "Rows": len(df), "Columns": len(df.columns)})

    for e in errors:
        st.markdown(f'<div class="error-box">{e}</div>', unsafe_allow_html=True)

    if not dfs:
        return

    total_rows = sum(m["Rows"] for m in file_meta)
    all_cols   = set()
    for df in dfs:
        all_cols.update(_normalize_col(c) for c in df.columns)

    st.markdown(f"""
    <div class="metric-row">
        <div class="metric-card"><div class="metric-value">{len(uploaded_files)}</div><div class="metric-label">Files Loaded</div></div>
        <div class="metric-card"><div class="metric-value">{len(dfs)}</div><div class="metric-label">Sheets Found</div></div>
        <div class="metric-card"><div class="metric-value">{total_rows:,}</div><div class="metric-label">Total Rows</div></div>
        <div class="metric-card"><div class="metric-value">{len(all_cols)}</div><div class="metric-label">Unique Columns</div></div>
    </div>
    """, unsafe_allow_html=True)

    with st.expander("\U0001f4cb Sheet Breakdown", expanded=True):
        st.dataframe(
            pd.DataFrame(file_meta),
            use_container_width=True, hide_index=True
        )

    st.markdown("<br>", unsafe_allow_html=True)
    col1, _, _ = st.columns([1, 1, 2])
    with col1:
        if st.button("Aggregate Files", key="agg_process"):
            st.session_state.pop("agg_output", None)
            st.session_state["agg_triggered"] = True

    if st.session_state.get("agg_triggered"):
        if "agg_output" not in st.session_state:
            with st.spinner("Aligning columns and merging files..."):
                try:
                    canon   = _build_canonical_columns(dfs)
                    aligned = [_align_df(df, canon) for df in dfs]
                    merged  = pd.concat(aligned, axis=0, ignore_index=True, sort=False)
                    first_cols = [canon[_normalize_col(c)] for c in dfs[0].columns if _normalize_col(c) in canon]
                    extra_cols = [c for c in merged.columns if c not in first_cols]
                    merged     = merged[first_cols + extra_cols]
                    st.session_state["agg_output"] = (_df_to_formatted_excel(merged, "Aggregated"), "Aggregated_Output.xlsx", None)
                    # Don't cache merged df — regenerate preview from bytes to save memory
                except Exception as exc:
                    st.session_state["agg_output"] = (None, None, str(exc))

        out_bytes, out_name, out_err = st.session_state["agg_output"]
        if out_err:
            st.markdown(f'<div class="error-box">{out_err}</div>', unsafe_allow_html=True)
        else:
            # Read preview from output bytes — single pass
            try:
                preview_df = pd.read_excel(io.BytesIO(out_bytes), sheet_name=0, nrows=20)
                full_idx   = pd.read_excel(io.BytesIO(out_bytes), sheet_name=0, usecols=[0])
                row_count  = len(full_idx); col_count = len(preview_df.columns); del full_idx
            except Exception:
                preview_df, row_count, col_count = pd.DataFrame(), 0, 0
            st.markdown(f'<div class="success-box">Aggregation complete &mdash; <b>{row_count:,} rows</b> across <b>{col_count} columns</b></div>', unsafe_allow_html=True)
            with st.expander("\U0001f4cb Merged Preview (first 20 rows)", expanded=True):
                st.dataframe(preview_df, use_container_width=True, height=300)
            col_a, _, _ = st.columns([1, 1, 2])
            with col_a:
                st.download_button(label="Download Aggregated File", data=out_bytes, file_name=out_name,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="agg_download")


# =============================================================================
# SUBTOTAL CALCULATOR TAB
# =============================================================================

def render_subtotal_tab():
    st.markdown("""
    <div style="margin-bottom:20px">
        <span style="font-family:\'IBM Plex Mono\',monospace;font-size:18px;font-weight:600;color:#80cbc4">\U0001f9ee  Subtotal Calculator</span>
        <span style="font-size:12px;color:#7a8aaa;margin-left:12px">Sum numeric columns across multiple files with identical structure</span>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="info-box"><b>How to use:</b><br>
    <span class="step-badge">1</span>Upload any number of Excel files that share the same column layout<br>
    <span class="step-badge">2</span>Click <b>Calculate Subtotals</b> &mdash; each file\'s numeric totals are computed<br>
    <span class="step-badge">3</span>Review the per-file breakdown and grand total row<br>
    <span class="step-badge">4</span>Download the summary as Excel
    </div>
    """, unsafe_allow_html=True)

    uploaded_files = st.file_uploader(
        "Upload files for subtotal calculation",
        type=["xlsx", "xls", "csv"],
        accept_multiple_files=True,
        key="sub_uploader",
        label_visibility="collapsed",
    )

    if not uploaded_files:
        st.markdown("""
        <div style="text-align:center;padding:48px 0;color:#3a4a6a;border:1.5px dashed #2a3a5c;border-radius:12px;margin-top:12px;">
            <div style="font-size:40px;margin-bottom:12px">&#x2191;</div>
            <div style="font-family:\'IBM Plex Mono\',monospace;font-size:14px">Drop your files here to calculate combined totals</div>
            <div style="font-size:12px;margin-top:6px">Excel (.xlsx, .xls) or CSV</div>
        </div>
        """, unsafe_allow_html=True)
        return

    dfs, names, errors = [], [], []
    for f in uploaded_files:
        df, err = _read_first_sheet(f)
        if err:
            errors.append(f"<b>{f.name}</b>: {err}")
        else:
            dfs.append(df)
            names.append(f.name)

    for e in errors:
        st.markdown(f'<div class="error-box">{e}</div>', unsafe_allow_html=True)

    if not dfs:
        return

    st.markdown(f"""
    <div class="metric-row">
        <div class="metric-card"><div class="metric-value">{len(dfs)}</div><div class="metric-label">Files Loaded</div></div>
        <div class="metric-card"><div class="metric-value">{sum(len(d) for d in dfs):,}</div><div class="metric-label">Total Rows</div></div>
    </div>
    """, unsafe_allow_html=True)

    col1, _, _ = st.columns([1, 1, 2])
    with col1:
        if st.button("Calculate Subtotals", key="sub_process"):
            st.session_state.pop("sub_output", None)
            st.session_state["sub_triggered"] = True

    if st.session_state.get("sub_triggered"):
        if "sub_output" not in st.session_state:
            with st.spinner("Calculating subtotals..."):
                try:
                    canon   = _build_canonical_columns(dfs)
                    aligned = [_align_df(df, canon) for df in dfs]

                    numeric_cols, seen_norm = [], set()
                    for df in aligned:
                        for col in df.columns:
                            n = _normalize_col(col)
                            if n not in seen_norm:
                                coerced = pd.to_numeric(df[col], errors="coerce")
                                if coerced.notna().any():
                                    numeric_cols.append(col)
                                    seen_norm.add(n)

                    if not numeric_cols:
                        st.session_state["sub_output"] = (None, None, "No numeric columns detected across the uploaded files.")
                    else:
                        rows = []
                        for name, df in zip(names, aligned):
                            row = {"File": name}
                            for col in numeric_cols:
                                if col in df.columns:
                                    row[col] = pd.to_numeric(df[col], errors="coerce").sum()
                                else:
                                    row[col] = 0.0
                            rows.append(row)
                        summary = pd.DataFrame(rows)
                        grand   = {"File": "GRAND TOTAL"}
                        for col in numeric_cols:
                            grand[col] = summary[col].sum()
                        summary = pd.concat([summary, pd.DataFrame([grand])], ignore_index=True)
                        st.session_state["sub_output"] = (
                            _df_to_formatted_excel(summary, "Subtotals"),
                            "Subtotals_Output.xlsx",
                            None,
                        )
                except Exception as exc:
                    st.session_state["sub_output"] = (None, None, str(exc))

        out_bytes, out_name, out_err = st.session_state["sub_output"]
        if out_err:
            st.markdown(f'<div class="error-box">{out_err}</div>', unsafe_allow_html=True)
        else:
            try:
                preview_df = pd.read_excel(io.BytesIO(out_bytes), sheet_name=0, nrows=50)
            except Exception:
                preview_df = pd.DataFrame()
            st.markdown(
                f'<div class="success-box">Subtotal calculation complete &mdash; '
                f'<b>{len(names)} file(s)</b> summarised</div>',
                unsafe_allow_html=True,
            )
            with st.expander("\U0001f4cb Subtotal Preview", expanded=True):
                st.dataframe(preview_df, use_container_width=True, height=300)
            col_a, _, _ = st.columns([1, 1, 2])
            with col_a:
                st.download_button(
                    label="Download Subtotals",
                    data=out_bytes,
                    file_name=out_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="sub_download",
                )


# =============================================================================
# MAIN — TAB LAYOUT
# =============================================================================

ALL_TABS = list(XDOCKS.keys()) + ["📁  Aggregator", "🧮  Subtotal"]
tabs = st.tabs(ALL_TABS)

for tab_idx, (xdock_display, xdock_cfg) in enumerate(XDOCKS.items()):
    with tabs[tab_idx]:
        xdock_key   = xdock_cfg["key"]
        xdock_color = xdock_cfg["color"]

        # ── Freezpak: inner Aggregator / Invoice Processing toggle ──────
        if xdock_key == "freezpak":
            st.markdown('<div class="mode-toggle-wrap">', unsafe_allow_html=True)
            fpk_mode = st.radio(
                "fpk_mode",
                ["📊  Aggregator", "🧾  Invoice Processing"],
                horizontal=True,
                key="fpk_mode",
                label_visibility="collapsed",
            )
            st.markdown('</div>', unsafe_allow_html=True)

            if fpk_mode == "📊  Aggregator":
                inv_types = INVOICE_TYPES.get("freezpak", [])
                inv_names = [it["name"] for it in inv_types]
                sel_name  = st.radio(
                    "Invoice Type",
                    inv_names,
                    horizontal=True,
                    key="inv_type_freezpak",
                    label_visibility="collapsed",
                )
                sel_cfg = next(it for it in inv_types if it["name"] == sel_name)
                render_invoice_section(xdock_key, sel_cfg, xdock_color, xdock_display)

            else:  # Invoice Processing
                proc_types = FREEZPAK_INVOICE_PROC_TYPES
                if len(proc_types) > 1:
                    proc_names   = [it["name"] for it in proc_types]
                    sel_proc     = st.radio(
                        "Invoice Type",
                        proc_names,
                        horizontal=True,
                        key="inv_proc_type_freezpak",
                        label_visibility="collapsed",
                    )
                    sel_cfg = next(it for it in proc_types if it["name"] == sel_proc)
                else:
                    sel_cfg = proc_types[0]
                render_invoice_section(xdock_key, sel_cfg, xdock_color, xdock_display)

        # ── All other XDock tabs ─────────────────────────────────────────
        else:
            inv_types = INVOICE_TYPES.get(xdock_key, [])
            if not inv_types:
                st.info("No invoice types configured for this XDock.")
                continue

            inv_names = [it["name"] for it in inv_types]
            sel_name  = st.radio(
                "Invoice Type",
                inv_names,
                horizontal=True,
                key=f"inv_type_{xdock_key}",
                label_visibility="collapsed",
            )
            sel_cfg = next(it for it in inv_types if it["name"] == sel_name)

            if "sub_types" in sel_cfg:
                sub_names    = [s["name"] for s in sel_cfg["sub_types"]]
                sel_sub_name = st.radio(
                    "Sub-Type",
                    sub_names,
                    horizontal=True,
                    key=f"sub_type_{xdock_key}",
                    label_visibility="collapsed",
                )
                sel_cfg = next(s for s in sel_cfg["sub_types"] if s["name"] == sel_sub_name)

            render_invoice_section(xdock_key, sel_cfg, xdock_color, xdock_display)

with tabs[len(XDOCKS)]:
    render_aggregator_tab()

with tabs[len(XDOCKS) + 1]:
    render_subtotal_tab()
