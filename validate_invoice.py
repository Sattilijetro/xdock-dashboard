"""
Validate invoice workbooks and export AP-only file when all rates pass.

Workflow:
  1. Paste Detail and AP tabs as values (no formulas)
  2. Rate-check Detail (AMOUNT / WEIGHT <= 0.475)
  3. Apply invoice suffixes (-1, -2, ... every 45 unique stores); keep all AP rows
  4. Save final workbook with AP tab only (Detail is not included)

Usage:
    python validate_invoice.py [path_to_workbook]

If no path is given, uses the newest full invoice .xlsx in the Template folder.
"""

from __future__ import annotations

import argparse
import re
import sys
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet

MAX_BILLABLE_RATE = 0.475
DETAIL_SHEET = "Detail"
AP_SHEET = "AP"
HEADER_ROW = 1
DATA_START_ROW = 2

COL_AMOUNT = 9   # I
COL_WEIGHT = 12  # L
COL_RATE_CHECK = 13  # M
COL_PO_NUMBER = 8  # H
COL_INVOICE = 3  # C
COL_WHS = 7  # G (warehouse used by AMOUNT array formula)
DETAIL_COL_PO_COMBINED = 6  # F (= WHS & PO NUMBER)
DETAIL_COL_PO_PART = 8  # H

DETAIL_REF_PATTERN = re.compile(r"=Detail!([A-Z]+)", re.IGNORECASE)

RATE_OK = "OK"
RATE_OVER = "Over Billable Rate"
RATE_CHECK_WEIGHT = "Check Weight"
RATE_CHECK_AMOUNT = "Check Amount"

# Billing rates per lb embedded in the AMOUNT (column I) array formula.
KNOWN_BILLING_RATES_PER_LB = (0.015, 0.01575, 0.04725)

WHS_RATE_PATTERN = re.compile(r"G\d+=(\d+),L\d+\*([\d.]+)")

STORES_PER_GROUP = 45
COL_AP_INVOICE_FALLBACK = 3  # C, used if invoice header is not found

PO_HEADER_CANDIDATES = [
    "PO Number",
    "PO Num",
    "PO_NUMBER",
    "P.O.",
    "P.O",
    "PO",
]
INVOICE_HEADER_CANDIDATES = [
    "INVOICE #",
    "INVOICE",
    "INV #",
    "INVOICE NUMBER",
]

INVOICE_IN_FILENAME = re.compile(r"invoice\s*(\d+)", re.IGNORECASE)


def _normalize_invoice_number(value) -> str:
    if value is None or value == "":
        raise ValueError("Invoice number is empty.")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0"):
        try:
            return str(int(float(text)))
        except ValueError:
            pass
    return text


def read_invoice_number(ws_values: Worksheet, workbook_path: Path) -> str:
    """Read invoice # from Detail column C (first data row), with filename fallback."""
    invoice = ws_values.cell(row=DATA_START_ROW, column=COL_INVOICE).value
    if invoice is not None and str(invoice).strip():
        return _normalize_invoice_number(invoice)

    match = INVOICE_IN_FILENAME.search(workbook_path.stem)
    if match:
        return match.group(1)

    raise ValueError(
        "Could not determine invoice number from Detail!C2 or the workbook filename."
    )


def output_ap_filename(workbook_path: Path) -> str:
    """Keep original workbook name with ' - AP' suffix."""
    return f"{workbook_path.stem} - AP.xlsx"


def output_exception_filename(workbook_path: Path) -> str:
    return f"{workbook_path.stem} - Exception Report.xlsx"


# Column layout for final AP export (matches standard AP upload format).
AP_EXPORT_COLUMN_WIDTHS = {
    "A": 13.140625,
    "B": 14.42578125,
    "C": 13.85546875,
    "D": 14.140625,
    "E": 9.28515625,
    "F": 13.7109375,
}
AP_EXPORT_LAST_COL = 6
PO_NUMBER_FORMAT = "0"
AMOUNT_NUMBER_FORMAT = '"$"#,##0.00'


def _normalize_header(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).lower())


def _header_matches(cell_value, candidate: str) -> bool:
    if cell_value is None:
        return False
    return _normalize_header(cell_value) == _normalize_header(candidate)


def find_ap_po_column(ws: Worksheet) -> int:
    """Locate PO Number column on AP by header name."""
    matches: list[tuple[int, int, str]] = []
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=HEADER_ROW, column=col).value
        if raw is None:
            continue
        raw_text = str(raw).strip()
        for priority, candidate in enumerate(PO_HEADER_CANDIDATES):
            if _header_matches(raw_text, candidate):
                matches.append((priority, col, raw_text))

    if not matches:
        raise ValueError(
            "Could not find PO Number column on AP tab. "
            f"Looked for headers like: {', '.join(PO_HEADER_CANDIDATES)}"
        )

    # Prefer higher-priority header names; break ties toward P.O.-style headers.
    matches.sort(key=lambda item: (item[0], 0 if "." in item[2] else 1, item[1]))
    return matches[0][1]


def find_ap_amount_column(ws: Worksheet) -> int | None:
    """Locate AMOUNT column on AP by header name."""
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=HEADER_ROW, column=col).value
        if raw is None:
            continue
        if _header_matches(raw, "AMOUNT"):
            return col
    return None


def find_ap_invoice_column(ws: Worksheet) -> int:
    """Locate invoice number column on AP by header name (fallback: column C)."""
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=HEADER_ROW, column=col).value
        if raw is None:
            continue
        for candidate in INVOICE_HEADER_CANDIDATES:
            if _header_matches(raw, candidate):
                return col

    return COL_AP_INVOICE_FALLBACK


def resolve_detail_po_number(ws_detail_values: Worksheet, row: int) -> str | None:
    """Build P.O. from Detail (F), or WHS (G) + PO NUMBER (H) when formulas are not cached."""
    po_number = ws_detail_values.cell(row=row, column=DETAIL_COL_PO_COMBINED).value
    if po_number is not None and str(po_number).strip() != "":
        return str(po_number).strip()

    whs = _normalize_whs(ws_detail_values.cell(row=row, column=COL_WHS).value)
    po_part = ws_detail_values.cell(row=row, column=DETAIL_COL_PO_PART).value
    if whs is None or po_part is None or str(po_part).strip() == "":
        return None

    po_part_text = str(int(po_part)) if isinstance(po_part, float) and po_part.is_integer() else str(po_part).strip()
    return f"{whs}{po_part_text}"


def apply_invoice_suffix_to_ap(
    target_ws: Worksheet,
    store_group_map: dict[str, str],
    original_invoice_number: str,
    po_col: int,
    invoice_col: int,
    data_start_row: int = DATA_START_ROW,
) -> None:
    """Update INVOICE # only; every row for the same store gets the same suffix."""
    for row in range(data_start_row, target_ws.max_row + 1):
        po_number = target_ws.cell(row=row, column=po_col).value
        if po_number is None or str(po_number).strip() == "":
            continue

        store_number = str(po_number).strip()[:3]
        suffix = store_group_map.get(store_number)
        if suffix is None:
            suffix = "-1"

        target_ws.cell(
            row=row,
            column=invoice_col,
            value=f"{original_invoice_number}{suffix}",
        )


def find_default_workbook(template_dir: Path) -> Path:
    """Newest full invoice workbook in Template (not AP-only exports)."""
    if not template_dir.is_dir():
        raise FileNotFoundError(f"Template folder not found: {template_dir}")

    candidates = [
        path
        for path in template_dir.glob("*.xlsx")
        if path.is_file()
        and not path.name.startswith("~$")
        and not path.stem.lower().endswith(" - ap")
    ]
    if not candidates:
        raise FileNotFoundError(
            f"No invoice workbook found in {template_dir}. "
            "Place your weekly .xlsx file there or pass the path as an argument."
        )

    return max(candidates, key=lambda path: path.stat().st_mtime)


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _is_blank_weight(weight) -> bool:
    if weight is None or weight == "":
        return True
    w = _to_float(weight)
    return w is None or w == 0


def _last_data_row(ws: Worksheet, start_row: int = DATA_START_ROW) -> int:
    last = HEADER_ROW
    for row in range(start_row, ws.max_row + 1):
        if any(
            ws.cell(row=row, column=col).value is not None
            for col in (COL_PO_NUMBER, COL_AMOUNT, COL_WEIGHT)
        ):
            last = row
    return last


def parse_whs_rate_map(formula_text: str) -> dict[str, float]:
    """Parse AMOUNT array IFS formula: warehouse -> rate per lb (e.g. 0.04725)."""
    rates: dict[str, float] = {}
    for whs, rate in WHS_RATE_PATTERN.findall(formula_text):
        rates[str(int(whs))] = float(rate)
    if re.search(r"G\d+=0,0", formula_text):
        rates["0"] = 0.0
    return rates


def get_whs_rate_map_from_detail(ws_formula: Worksheet) -> dict[str, float]:
    """Read warehouse rate table from the AMOUNT column array formula."""
    for row in range(DATA_START_ROW, min(ws_formula.max_row, DATA_START_ROW + 5) + 1):
        value = ws_formula.cell(row=row, column=COL_AMOUNT).value
        if isinstance(value, ArrayFormula):
            return parse_whs_rate_map(value.text)
        if isinstance(value, str) and value.lstrip().upper().startswith("=IFS"):
            return parse_whs_rate_map(value)
    return {}


def _normalize_whs(value) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0"):
        try:
            return str(int(float(text)))
        except ValueError:
            pass
    return text


def resolve_detail_amount_and_rate(
    ws_values: Worksheet,
    row: int,
    whs_rate_map: dict[str, float],
) -> tuple[float | None, float | None]:
    """
    Resolve AMOUNT and AMOUNT/WEIGHT rate for a Detail row.

    Column I often uses an array formula (WEIGHT * rate by warehouse). When Excel
    cached values are missing, derive amount as WEIGHT * rate using column G (WHS).
    """
    weight = ws_values.cell(row=row, column=COL_WEIGHT).value
    if _is_blank_weight(weight):
        return None, None

    wgt = _to_float(weight)
    if wgt is None:
        return None, None

    amount = _to_float(ws_values.cell(row=row, column=COL_AMOUNT).value)
    if amount is not None:
        return amount, amount / wgt

    whs_key = _normalize_whs(ws_values.cell(row=row, column=COL_WHS).value)
    if whs_key is None or whs_key not in whs_rate_map:
        return None, None

    rate_per_lb = whs_rate_map[whs_key]
    return wgt * rate_per_lb, rate_per_lb


def _classify_row(rate: float | None, weight) -> tuple[str, float | None]:
    if _is_blank_weight(weight):
        return RATE_CHECK_WEIGHT, None

    if rate is None:
        return RATE_CHECK_AMOUNT, None

    if rate > MAX_BILLABLE_RATE:
        return RATE_OVER, rate
    return RATE_OK, rate


def _is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _paste_detail_sheet_as_values(
    ws: Worksheet, ws_cached: Worksheet, whs_rate_map: dict[str, float]
) -> None:
    """Replace Detail formulas with calculated values (paste as values)."""
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            value = cell.value
            if value is None:
                continue

            if isinstance(value, ArrayFormula) and col == COL_AMOUNT:
                amount, _ = resolve_detail_amount_and_rate(ws_cached, row, whs_rate_map)
                cell.value = amount
            elif _is_formula(value):
                if col == DETAIL_COL_PO_COMBINED:
                    cell.value = resolve_detail_po_number(ws_cached, row)
                else:
                    cached = ws_cached.cell(row=row, column=col).value
                    if cached is not None:
                        cell.value = cached
            elif isinstance(value, ArrayFormula):
                cached = ws_cached.cell(row=row, column=col).value
                if cached is not None:
                    cell.value = cached


def _paste_ap_sheet_as_values(ws_ap: Worksheet, ws_cached: Worksheet, ws_detail: Worksheet) -> None:
    """Replace AP formulas with values (Detail links read pasted Detail values)."""
    for row in range(1, ws_ap.max_row + 1):
        for col in range(1, ws_ap.max_column + 1):
            cell = ws_ap.cell(row=row, column=col)
            value = cell.value
            if value is None:
                continue

            if _is_formula(value):
                match = DETAIL_REF_PATTERN.match(str(value).strip())
                if match:
                    detail_col = column_index_from_string(match.group(1))
                    cell.value = ws_detail.cell(row=row, column=detail_col).value
                else:
                    cached = ws_cached.cell(row=row, column=col).value
                    if cached is not None:
                        cell.value = cached


def paste_workbook_as_values(wb, wb_cached) -> None:
    """Step 1: Convert Detail and AP tabs to values before validation/export."""
    whs_rate_map = get_whs_rate_map_from_detail(wb[DETAIL_SHEET])
    _paste_detail_sheet_as_values(wb[DETAIL_SHEET], wb_cached[DETAIL_SHEET], whs_rate_map)
    _paste_ap_sheet_as_values(wb[AP_SHEET], wb_cached[AP_SHEET], wb[DETAIL_SHEET])


def _ap_po_value(ws: Worksheet, row: int, po_col: int) -> str | None:
    po_number = ws.cell(row=row, column=po_col).value
    if po_number is None or str(po_number).strip() == "":
        return None
    if isinstance(po_number, float) and po_number.is_integer():
        return str(int(po_number))
    return str(po_number).strip()


def _last_ap_data_row_values(ws: Worksheet, po_col: int) -> int:
    last = HEADER_ROW
    for row in range(DATA_START_ROW, ws.max_row + 1):
        if _ap_po_value(ws, row, po_col):
            last = row
    return last


def get_store_group_map_from_ap(ws_ap_values: Worksheet, po_col: int, last_row: int) -> dict[str, str]:
    """
    Build store -> invoice suffix map from AP rows.

    Each unique store (first 3 digits of P.O.) counts once toward the 45-store
    groups, even if that store appears on many rows. All rows for store 895 share
    the same suffix; rows are not removed from the export.
    """
    unique_stores: list[str] = []
    seen_stores: set[str] = set()

    for row in range(DATA_START_ROW, last_row + 1):
        po_number = _ap_po_value(ws_ap_values, row, po_col)
        if po_number is None:
            continue

        store_number = po_number[:3]
        if store_number in seen_stores:
            continue

        seen_stores.add(store_number)
        unique_stores.append(store_number)

    store_group_map: dict[str, str] = {}
    for index, store_number in enumerate(unique_stores):
        group_number = index // STORES_PER_GROUP + 1
        store_group_map[store_number] = f"-{group_number}"

    return store_group_map


def _all_ap_data_rows(last_row: int) -> list[int]:
    """All AP data row numbers (no deduplication by store)."""
    return list(range(DATA_START_ROW, last_row + 1))


def validate_detail(
    ws_formula: Worksheet,
    ws_values: Worksheet,
    ws_write: Worksheet,
) -> list[dict]:
    """Update Rate Check (column M) and return exception rows."""
    exceptions: list[dict] = []
    whs_rate_map = get_whs_rate_map_from_detail(ws_formula)

    if ws_write.cell(HEADER_ROW, COL_RATE_CHECK).value != "Rate Check":
        ws_write.cell(HEADER_ROW, COL_RATE_CHECK, value="Rate Check")

    last_row = _last_data_row(ws_values)
    for row in range(DATA_START_ROW, last_row + 1):
        weight = ws_values.cell(row=row, column=COL_WEIGHT).value
        po_number = ws_values.cell(row=row, column=COL_PO_NUMBER).value
        amount, rate = resolve_detail_amount_and_rate(ws_values, row, whs_rate_map)

        status, rate = _classify_row(rate, weight)
        ws_write.cell(row=row, column=COL_RATE_CHECK, value=status)

        if status != RATE_OK:
            exceptions.append(
                {
                    "row_number": row,
                    "po_number": po_number,
                    "amount": amount,
                    "weight": weight,
                    "calculated_rate": rate,
                    "issue": status,
                }
            )

    return exceptions


def _copy_sheet_layout(ws_formula: Worksheet, target_ws: Worksheet) -> None:
    target_ws.sheet_format = copy(ws_formula.sheet_format)
    target_ws.sheet_properties = copy(ws_formula.sheet_properties)
    target_ws.merged_cells = copy(ws_formula.merged_cells)
    target_ws.page_setup = copy(ws_formula.page_setup)
    target_ws.print_options = copy(ws_formula.print_options)

    for col_letter, dim in ws_formula.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dim.width
        target_ws.column_dimensions[col_letter].hidden = dim.hidden
        if dim.bestFit is not None:
            target_ws.column_dimensions[col_letter].bestFit = dim.bestFit

    for row_idx, dim in ws_formula.row_dimensions.items():
        target_ws.row_dimensions[row_idx].height = dim.height
        target_ws.row_dimensions[row_idx].hidden = dim.hidden


def _to_po_number(value):
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip().replace(",", "")
    return int(float(text))


def format_ap_export_sheet(
    ws: Worksheet,
    po_col: int,
    amount_col: int | None,
    last_row: int,
    last_col: int = AP_EXPORT_LAST_COL,
) -> None:
    """Apply PO number format, currency on AMOUNT, column widths, and header filter."""
    for letter, width in AP_EXPORT_COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    for row in range(DATA_START_ROW, last_row + 1):
        po_cell = ws.cell(row=row, column=po_col)
        po_cell.value = _to_po_number(po_cell.value)
        po_cell.number_format = PO_NUMBER_FORMAT

        if amount_col is not None:
            amount_cell = ws.cell(row=row, column=amount_col)
            if amount_cell.value is not None:
                amount_cell.value = float(amount_cell.value)
                amount_cell.number_format = AMOUNT_NUMBER_FORMAT

    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"


def _copy_row_values_and_styles(
    ws_source: Worksheet,
    source_row: int,
    target_ws: Worksheet,
    target_row: int,
    last_col: int,
) -> None:
    for col in range(1, last_col + 1):
        source_cell = ws_source.cell(row=source_row, column=col)
        target_cell = target_ws.cell(row=target_row, column=col)
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)


def build_ap_export_workbook(
    ws_ap: Worksheet,
    output_path: Path,
    original_invoice_number: str,
) -> tuple[int, int]:
    """
    Export final AP-only workbook (no Detail tab).
    Keeps every AP row; only INVOICE # is updated with store-based suffixes.
    """
    po_col = find_ap_po_column(ws_ap)
    invoice_col = find_ap_invoice_column(ws_ap)
    last_ap_row = _last_ap_data_row_values(ws_ap, po_col)

    if last_ap_row < DATA_START_ROW:
        raise ValueError(
            "No AP rows could be exported. P.O. values may be missing on the AP tab."
        )

    store_group_map = get_store_group_map_from_ap(ws_ap, po_col, last_ap_row)
    all_rows = _all_ap_data_rows(last_ap_row)

    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    target_ws = out_wb.create_sheet(title=AP_SHEET)
    last_col = AP_EXPORT_LAST_COL

    _copy_sheet_layout(ws_ap, target_ws)

    target_row = HEADER_ROW
    for source_row in [HEADER_ROW, *all_rows]:
        _copy_row_values_and_styles(ws_ap, source_row, target_ws, target_row, last_col)
        target_row += 1

    last_export_row = target_row - 1

    apply_invoice_suffix_to_ap(
        target_ws,
        store_group_map,
        original_invoice_number,
        po_col,
        invoice_col,
    )

    amount_col = find_ap_amount_column(target_ws)
    format_ap_export_sheet(
        target_ws, po_col, amount_col, last_export_row, last_col
    )

    unique_store_count = len(store_group_map)
    suffixes = sorted(set(store_group_map.values()))
    print(
        f"  Invoice suffixes in column {invoice_col} "
        f"({ws_ap.cell(HEADER_ROW, invoice_col).value}): "
        f"{', '.join(original_invoice_number + s for s in suffixes)}"
    )
    print(
        f"  {unique_store_count} unique store(s) across {len(all_rows)} AP row(s) "
        f"(repeat stores share the same suffix)."
    )

    out_wb.save(output_path)
    return len(all_rows), unique_store_count


def write_exception_report(exceptions: list[dict], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Exceptions"
    headers = [
        "Row Number",
        "PO Number",
        "Amount",
        "Weight",
        "Calculated Rate",
        "Issue",
    ]
    ws.append(headers)
    for item in exceptions:
        ws.append(
            [
                item["row_number"],
                item["po_number"],
                item["amount"],
                item["weight"],
                item["calculated_rate"],
                item["issue"],
            ]
        )
    wb.save(output_path)


def run(workbook_path: Path, output_dir: Path) -> int:
    if not workbook_path.is_file():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 1

    wb = load_workbook(workbook_path, data_only=False)
    wb_cached = load_workbook(workbook_path, data_only=True)

    if DETAIL_SHEET not in wb.sheetnames or AP_SHEET not in wb.sheetnames:
        print(f"Expected sheets '{AP_SHEET}' and '{DETAIL_SHEET}'.", file=sys.stderr)
        return 1

    print("Step 1: Pasting Detail and AP as values...")
    paste_workbook_as_values(wb, wb_cached)

    ws_detail = wb[DETAIL_SHEET]
    try:
        invoice_number = read_invoice_number(ws_detail, workbook_path)
    except ValueError as exc:
        print(exc, file=sys.stderr)
        return 1

    print(f"Invoice number: {invoice_number}")
    print("Step 2: Rate check on Detail tab...")
    exceptions = validate_detail(ws_detail, ws_detail, ws_detail)
    last_detail_row = _last_data_row(ws_detail)

    output_dir.mkdir(parents=True, exist_ok=True)

    if exceptions:
        exception_path = output_dir / output_exception_filename(workbook_path)
        write_exception_report(exceptions, exception_path)
        print(
            f"Validation failed: {len(exceptions)} row(s) need review. "
            f"Exception report saved to {exception_path}"
        )
        print("AP tab was not exported.")
        return 2

    print("Step 3: Building AP-only export (invoice suffixes by unique store)...")
    ap_output = output_dir / output_ap_filename(workbook_path)
    try:
        ap_rows_exported, unique_stores = build_ap_export_workbook(
            wb[AP_SHEET], ap_output, invoice_number
        )
    except ValueError as exc:
        print(exc, file=sys.stderr)
        return 1

    suffix_groups = (unique_stores + STORES_PER_GROUP - 1) // STORES_PER_GROUP
    print(
        f"All {last_detail_row - HEADER_ROW} detail row(s) passed rate validation.\n"
        f"Final file (AP tab only, no Detail): {ap_output}\n"
        f"  {ap_rows_exported} AP row(s), {unique_stores} unique store(s), "
        f"{suffix_groups} invoice suffix group(s)."
    )
    return 0


def main() -> int:
    base_dir = Path(__file__).resolve().parent
    template_dir = base_dir / "Template"

    parser = argparse.ArgumentParser(description="Validate invoice rates and export AP.")
    parser.add_argument(
        "workbook",
        nargs="?",
        default=None,
        help="Path to the invoice workbook (.xlsx). "
        "If omitted, uses the newest invoice in the Template folder.",
    )
    parser.add_argument(
        "-o",
        "--output-dir",
        default=str(base_dir),
        help="Directory for output file(s)",
    )
    args = parser.parse_args()

    if args.workbook:
        workbook_path = Path(args.workbook)
    else:
        try:
            workbook_path = find_default_workbook(template_dir)
            print(f"Using workbook: {workbook_path}")
        except FileNotFoundError as exc:
            print(exc, file=sys.stderr)
            return 1

    return run(workbook_path, Path(args.output_dir))


if __name__ == "__main__":
    raise SystemExit(main())
